import tkinter as tk
from tkinter import messagebox
import openpyxl

class tkinter:
    def __init__(self, master):
        self.master = master
        self.master.title("Generador de Informes de Gastos")
        
        self.g = []

        self.lf = tk.Label(master, text="Fecha (DD/MM/AAAA):")
        self.ef = tk.Entry(master)
        self.ld = tk.Label(master, text="Descripción:")
        self.ed = tk.Entry(master)
        self.lm = tk.Label(master, text="Monto:")
        self.em = tk.Entry(master)

       
        self.ba = tk.Button(master, text="Agregar Gasto", command=self.agregar)

        
        self.br = tk.Button(master, text="Mostrar Resumen", command=self.imprimir)

        
        self.lf.grid(row=0, column=0, padx=5, pady=5)
        self.ef.grid(row=0, column=1, padx=5, pady=5)
        self.ld.grid(row=1, column=0, padx=5, pady=5)
        self.ed.grid(row=1, column=1, padx=5, pady=5)
        self.lm.grid(row=2, column=0, padx=5, pady=5)
        self.em.grid(row=2, column=1, padx=5, pady=5)
        self.ba.grid(row=3, column=0, columnspan=2, pady=10)
        self.br.grid(row=4, column=0, columnspan=2, pady=10)

        
        self.ln = tk.Label(master, text="Melvin Daniel Sandoval Mejía")
        self.ln.grid(row=5, column=0, columnspan=2, pady=10)
        

    def agregar(self):
        fecha = self.ef.get()
        descripcion = self.ed.get()
        monto = self.em.get()

        if fecha and descripcion and monto:
            try:
                monto = float(monto)
                gasto = {'fecha': fecha, 'descripcion': descripcion, 'monto': monto}
                self.g.append(gasto)

                
                self.ef.delete(0, tk.END)
                self.ed.delete(0, tk.END)
                self.em.delete(0, tk.END)
            except ValueError:
                messagebox.showerror("Error", "Por favor, ingrese un monto válido.")
        else:
            messagebox.showwarning("Advertencia", "Por favor, complete todos los campos.")

    def imprimir(self):
        if not self.g:
            messagebox.showwarning("Advertencia", "No hay gastos para mostrar.")
            return

        resumen = self.calculo()

        messagebox.showinfo("Resumen",
                            f"Número de Gastos: {resumen[0]}\n"
                            f"Fecha y Descripción del Gasto Más Caro: {resumen[1]['fecha']} - {resumen[1]['descripcion']}\n"
                            f"Fecha y Descripción del Gasto Más Barato: {resumen[2]['fecha']} - {resumen[2]['descripcion']}\n"
                            f"Monto Total de Gastos: ${resumen[3]:,.2f}")

        self.guardar(resumen)

    def calculo(self):
        numeracion = len(self.g)
        total = sum(g['monto'] for g in self.g)

        caro = max(self.g, key=lambda x: x['monto'])
        barato = min(self.g, key=lambda x: x['monto'])

        return numeracion, caro, barato, total

    def guardar(self, resumen):
        wb = openpyxl.Workbook()
        
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        sheet = wb.create_sheet(title="Gastos")

        
        sheet.append(["Fecha", "Descripción", "Monto"])

        for g in self.g:
            sheet.append([g['fecha'], g['descripcion'], g['monto']])

        sheet.append([])  
        sheet.append(["Número de Gastos", "Fecha del Gasto Más Caro", "Descripción del Gasto Más Caro",
                      "Fecha del Gasto Más Barato", "Descripción del Gasto Más Barato", "Total de Gastos"])

        numeracion, caro, barato, total = resumen
        sheet.append([numeracion, caro['fecha'], caro['descripcion'],
                      barato['fecha'], barato['descripcion'], total])

        wb.save("C:/Users/Florecita/OneDrive/Documentos/Python/Openpyxl/informe_gastos.xlsx")


def main():
    root = tk.Tk()
    app = tkinter(root)
    root.mainloop()


if __name__ == "__main__":
    main()





